"""
魔方原声处理 v1.1.0
功能：解析咨询CSV和聊天记录LOG，生成分析表格，支持写入飞书和导出Excel
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import zipfile
import re
import os
import sys
import csv
import io
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import requests
import threading
import configparser
from urllib.parse import urlparse, parse_qs
from collections import defaultdict, OrderedDict


VERSION = "1.1.0"
APP_TITLE = f"魔方原声处理 v{VERSION}"
APP_DATA_DIR = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "魔方原声处理")
CONFIG_FILE = os.path.join(APP_DATA_DIR, "feishu_config.ini")
PURCHASE_KEYWORD_DEFAULT = "已收到你的订单"
_RE_URL = re.compile(r"https?://\S+", re.IGNORECASE)
CONSULT_EXTS = (".zip", ".xlsx")
LOG_EXTS = (".log",)

# 输出列顺序
OUTPUT_COLUMNS = ["咨询时间", "客户ID", "接待客服", "商品编号", "进线次数", "是否购买", "聊天记录", "客户问题原话", "客服回复"]


def resource_path(name):
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, name)


class ChatAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("980x760")
        self.root.minsize(900, 680)
        self.root.resizable(True, True)

        # 设置窗口图标
        icon_path = resource_path("bee_icon.ico")
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass

        self.csv_data = []
        self.chat_sessions = defaultdict(list)
        self.result_data = []

        self.zip_path = tk.StringVar()
        self.log_path = tk.StringVar()
        self.feishu_app_id = tk.StringVar()
        self.feishu_app_secret = tk.StringVar()
        self.feishu_link = tk.StringVar()
        self.feishu_create_new = tk.BooleanVar(value=False)
        self.feishu_create_new_btn = None
        self.feishu_table_name = tk.StringVar(value="原声分析")
        self.status_text = tk.StringVar(value="就绪")
        self.stat_csv = tk.StringVar(value="0")
        self.stat_sessions = tk.StringVar(value="0")
        self.stat_output = tk.StringVar(value="0")
        self.last_output_dir = os.path.abspath(os.path.dirname(__file__))
        self.is_busy = False
        self.action_buttons = []

        self._load_config()
        self._build_ui()
        self._load_keywords_config()

    # ───────────────── UI ─────────────────
    def _build_ui(self):
        self.root.geometry("980x760")
        self.root.minsize(900, 680)
        self.root.configure(bg="#F3F6FA")

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        base_font = ("Microsoft YaHei UI", 10)
        title_font = ("Microsoft YaHei UI", 18, "bold")
        style.configure(".", font=base_font)
        style.configure("TFrame", background="#F3F6FA")
        style.configure("Card.TFrame", background="#FFFFFF")
        style.configure("Sidebar.TFrame", background="#172033")
        style.configure("TLabel", background="#F3F6FA", foreground="#1F2937")
        style.configure("Card.TLabel", background="#FFFFFF", foreground="#1F2937")
        style.configure("Muted.TLabel", background="#FFFFFF", foreground="#6B7280", font=("Microsoft YaHei UI", 9))
        style.configure("HeaderMuted.TLabel", background="#F3F6FA", foreground="#6B7280")
        style.configure("SidebarTitle.TLabel", background="#172033", foreground="#FFFFFF", font=("Microsoft YaHei UI", 14, "bold"))
        style.configure("SidebarItem.TLabel", background="#172033", foreground="#CBD5E1", font=("Microsoft YaHei UI", 10))
        style.configure("StatValue.TLabel", background="#FFFFFF", foreground="#2563EB", font=("Microsoft YaHei UI", 18, "bold"))
        style.configure("TButton", padding=(8, 4), background="#FFFFFF")
        style.configure("Small.TButton", padding=(6, 3), font=("Microsoft YaHei UI", 9))
        style.configure("Primary.TButton", padding=(12, 6), background="#2563EB", foreground="#FFFFFF", font=("Microsoft YaHei UI", 10, "bold"))
        style.map("Primary.TButton", background=[("active", "#1D4ED8"), ("disabled", "#93C5FD")], foreground=[("disabled", "#E5E7EB")])
        style.configure("Accent.TButton", padding=(14, 7), background="#EAF1FF", foreground="#1D4ED8")
        style.configure("Horizontal.TProgressbar", troughcolor="#E5E7EB", background="#2563EB", bordercolor="#E5E7EB", lightcolor="#2563EB", darkcolor="#2563EB")

        shell = ttk.Frame(self.root, padding=(14, 12, 14, 12))
        shell.pack(fill="both", expand=True)
        shell.columnconfigure(0, weight=1)
        shell.rowconfigure(0, weight=0)
        shell.rowconfigure(1, weight=2)

        main = ttk.Frame(shell)
        main.grid(row=0, column=0, sticky="ew")
        main.columnconfigure(0, weight=1)

        content = ttk.Frame(main)
        content.grid(row=0, column=0, sticky="ew")
        content.columnconfigure(0, weight=1)

        feishu_card = self._create_card(content, "飞书")
        feishu_card.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        feishu_card.columnconfigure(1, weight=1)
        feishu_card.columnconfigure(3, weight=1)
        ttk.Label(feishu_card, text="App ID", style="Card.TLabel").grid(row=1, column=0, sticky="w", pady=4)
        ttk.Entry(feishu_card, textvariable=self.feishu_app_id).grid(row=1, column=1, sticky="ew", padx=(10, 18), pady=4)
        ttk.Label(feishu_card, text="App Secret", style="Card.TLabel").grid(row=1, column=2, sticky="w", pady=4)
        ttk.Entry(feishu_card, textvariable=self.feishu_app_secret, show="*").grid(row=1, column=3, sticky="ew", padx=(10, 0), pady=4)
        ttk.Label(feishu_card, text="多维表格链接", style="Card.TLabel").grid(row=2, column=0, sticky="w", pady=4)
        ttk.Entry(feishu_card, textvariable=self.feishu_link).grid(row=2, column=1, columnspan=3, sticky="ew", padx=(10, 0), pady=4)
        fs_actions = ttk.Frame(feishu_card, style="Card.TFrame")
        fs_actions.grid(row=3, column=1, columnspan=3, sticky="w", pady=(4, 0))
        self.feishu_create_new_btn = tk.Button(
            fs_actions, text="", command=self._toggle_feishu_create_new,
            relief="flat", bd=0, cursor="hand2", bg="#FFFFFF", fg="#1F2937",
            activebackground="#EFF6FF", activeforeground="#1D4ED8",
            font=("Segoe UI Emoji", 10), padx=2, pady=2
        )
        self.feishu_create_new_btn.pack(side="left", padx=(0, 14))
        self._sync_feishu_create_new_label()
        ttk.Label(fs_actions, text="表名", style="Card.TLabel").pack(side="left", padx=(0, 6))
        ttk.Entry(fs_actions, textvariable=self.feishu_table_name, width=18).pack(side="left", padx=(0, 14))
        ttk.Button(fs_actions, text="测试连接", command=self._on_test_feishu, style="Small.TButton", width=10).pack(side="left", padx=(0, 8))
        ttk.Button(fs_actions, text="保存配置", command=self._on_save_config, style="Small.TButton", width=10).pack(side="left")

        input_card = self._create_card(content, "数据")
        input_card.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        input_card.columnconfigure(1, weight=1)
        self._create_path_picker(input_card, 1, "咨询数据", self.zip_path,
                                 self._select_consult_file, self._select_consult_folder,
                                 "支持 .zip / .xlsx，也可以选择文件夹批量导入")
        self._create_path_picker(input_card, 3, "聊天记录", self.log_path,
                                 self._select_log_file, self._select_log_folder,
                                 "支持 .log，也可以选择文件夹批量导入")

        setting_card = self._create_card(content, "筛选")
        setting_card.grid(row=2, column=0, sticky="ew", pady=(0, 8))
        setting_card.columnconfigure(0, weight=1)
        setting_card.columnconfigure(1, weight=1)
        ttk.Label(setting_card, text="SKU筛选", style="Card.TLabel").grid(row=1, column=0, sticky="w", pady=(2, 4), padx=(0, 8))
        ttk.Label(setting_card, text="成交话术", style="Card.TLabel").grid(row=1, column=1, sticky="w", pady=(2, 4), padx=(8, 0))

        sku_frame = ttk.Frame(setting_card, style="Card.TFrame")
        sku_frame.grid(row=2, column=0, sticky="ew", padx=(0, 8))
        sku_frame.columnconfigure(0, weight=1)
        self.sku_text = tk.Text(sku_frame, height=3, wrap="word", font=("Microsoft YaHei UI", 9),
                                relief="flat", borderwidth=0, bg="#F8FAFC", fg="#111827",
                                padx=8, pady=6, highlightthickness=1,
                                highlightbackground="#CBD5E1", highlightcolor="#2563EB")
        self.sku_text.grid(row=0, column=0, sticky="ew")
        ttk.Label(setting_card, text="每行一个SKU，留空输出全部", style="Muted.TLabel").grid(row=3, column=0, sticky="w", pady=(2, 0), padx=(0, 8))

        kw_frame = ttk.Frame(setting_card, style="Card.TFrame")
        kw_frame.grid(row=2, column=1, sticky="ew", padx=(8, 0))
        kw_frame.columnconfigure(0, weight=1)
        self.keyword_text = tk.Text(kw_frame, height=3, wrap="word", font=("Microsoft YaHei UI", 9),
                                    relief="flat", borderwidth=0, bg="#F8FAFC", fg="#111827",
                                    padx=8, pady=6, highlightthickness=1,
                                    highlightbackground="#CBD5E1", highlightcolor="#2563EB")
        self.keyword_text.grid(row=0, column=0, sticky="ew")
        self.keyword_text.insert("1.0", PURCHASE_KEYWORD_DEFAULT)
        ttk.Label(setting_card, text="每行一条，任意匹配即已购买", style="Muted.TLabel").grid(row=3, column=1, sticky="w", pady=(2, 0), padx=(8, 0))

        action_card = self._create_card(content, "操作区")
        action_card.grid(row=3, column=0, sticky="ew", pady=(0, 8))
        action_card.columnconfigure((0, 1, 2, 3), weight=1)
        self.btn_parse = ttk.Button(action_card, text="开始解析", command=self._on_parse, style="Primary.TButton")
        self.btn_feishu = ttk.Button(action_card, text="写入飞书", command=self._on_write_feishu)
        self.btn_excel = ttk.Button(action_card, text="生成Excel", command=self._on_export_excel)
        self.btn_parse.grid(row=1, column=0, sticky="ew", padx=(0, 8), pady=(2, 6))
        self.btn_feishu.grid(row=1, column=1, sticky="ew", padx=4, pady=(2, 6))
        self.btn_excel.grid(row=1, column=2, sticky="ew", padx=4, pady=(2, 6))
        self.action_buttons = [self.btn_parse, self.btn_feishu, self.btn_excel]
        ttk.Button(action_card, text="打开输出目录", command=self._open_output_dir).grid(row=1, column=3, sticky="ew", padx=(8, 0), pady=(2, 6))

        stats = ttk.Frame(action_card, style="Card.TFrame")
        stats.grid(row=2, column=0, columnspan=4, sticky="ew", pady=(0, 0))
        stats.columnconfigure((0, 1, 2), weight=1)
        self._create_stat(stats, 0, "CSV记录数", self.stat_csv)
        self._create_stat(stats, 1, "聊天会话数", self.stat_sessions)
        self._create_stat(stats, 2, "输出行数", self.stat_output)

        log_container = ttk.Frame(shell)
        log_container.grid(row=1, column=0, sticky="nsew", pady=(0, 0))
        log_container.columnconfigure(0, weight=1)
        log_container.rowconfigure(0, weight=1)
        log_card = self._create_card(main, "运行日志")
        log_card.grid_forget()
        log_card = self._create_card(log_container, "运行日志")
        log_card.grid(row=0, column=0, sticky="nsew")
        log_card.columnconfigure(0, weight=1)
        log_card.rowconfigure(1, weight=1)
        self.progress = ttk.Progressbar(log_card, mode="determinate")
        log_wrap = ttk.Frame(log_card, style="Card.TFrame")
        log_wrap.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        log_wrap.columnconfigure(0, weight=1)
        log_wrap.rowconfigure(0, weight=1)
        self.log_text = tk.Text(log_wrap, height=11, wrap="word", font=("Consolas", 9),
                                relief="flat", borderwidth=0, bg="#0F172A", fg="#E5E7EB",
                                insertbackground="#E5E7EB")
        self.log_text.grid(row=0, column=0, sticky="nsew")

    def _create_card(self, parent, title, subtitle=None):
        card = tk.Frame(parent, bg="#FFFFFF", padx=12, pady=8,
                        highlightthickness=1, highlightbackground="#DCE3ED", highlightcolor="#DCE3ED")
        ttk.Label(card, text=title, style="Card.TLabel",
                  font=("Microsoft YaHei UI", 10, "bold")).grid(row=0, column=0, columnspan=4, sticky="w")
        if subtitle:
            ttk.Label(card, text=subtitle, style="Muted.TLabel", wraplength=220).grid(row=0, column=1, columnspan=3, sticky="e", padx=(12, 0))
        return card

    def _create_path_picker(self, parent, row, label, variable, file_command, folder_command, hint):
        ttk.Label(parent, text=label, style="Card.TLabel").grid(row=row, column=0, sticky="w", pady=(4, 2))
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", padx=(10, 8), pady=(4, 2))
        ttk.Button(parent, text="选文件", command=file_command, style="Small.TButton", width=8).grid(row=row, column=2, sticky="ew", padx=(0, 6), pady=(4, 2))
        ttk.Button(parent, text="选文件夹", command=folder_command, style="Small.TButton", width=9).grid(row=row, column=3, sticky="ew", pady=(4, 2))
        ttk.Label(parent, text=hint, style="Muted.TLabel", wraplength=520).grid(row=row + 1, column=1, columnspan=3, sticky="w", padx=(10, 0), pady=(0, 2))

    def _create_stat(self, parent, col, title, variable):
        box = tk.Frame(parent, bg="#F8FAFC", padx=8, pady=2,
                       highlightthickness=1, highlightbackground="#E5EAF2")
        box.grid(row=0, column=col, sticky="ew", padx=(0 if col == 0 else 6, 0 if col == 2 else 6))
        tk.Label(box, text=title, bg="#F8FAFC", fg="#64748B", font=("Microsoft YaHei UI", 8)).pack(anchor="center")
        tk.Label(box, textvariable=variable, bg="#F8FAFC", fg="#2563EB", font=("Microsoft YaHei UI", 11, "bold")).pack(anchor="center")

    def _toggle_feishu_create_new(self):
        self.feishu_create_new.set(not self.feishu_create_new.get())
        self._sync_feishu_create_new_label()

    def _sync_feishu_create_new_label(self):
        if not self.feishu_create_new_btn:
            return
        checked = self.feishu_create_new.get()
        self.feishu_create_new_btn.configure(
            text=("✅ 新建表格" if checked else "☐ 新建表格"),
            fg=("#047857" if checked else "#1F2937")
        )

    def _update_stats(self, csv_count=0, session_count=0, output_count=0):
        self.stat_csv.set(str(csv_count))
        self.stat_sessions.set(str(session_count))
        self.stat_output.set(str(output_count))

    def _clear_log(self):
        self.log_text.delete("1.0", "end")
        self.progress["value"] = 0

    def _open_output_dir(self):
        target = self.last_output_dir if os.path.isdir(self.last_output_dir) else os.getcwd()
        try:
            os.startfile(target)
        except Exception as e:
            messagebox.showerror("错误", f"无法打开输出目录:\n{e}")

    # ───────────────── 辅助 ─────────────────
    def _log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{ts}] {msg}\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def _set_busy(self, busy, status=None):
        self.is_busy = busy
        for btn in self.action_buttons:
            btn.configure(state="disabled" if busy else "normal")
        self.status_text.set(status or ("处理中..." if busy else "就绪"))

    def _start_task(self, status, target):
        if self.is_busy:
            messagebox.showwarning("提示", "当前任务还在执行，请稍等完成后再操作")
            return
        self._set_busy(True, status)
        threading.Thread(target=target, daemon=True).start()

    def _finish_task(self):
        self._set_busy(False, "就绪")

    def _select_consult_file(self):
        p = filedialog.askopenfilename(title="选择咨询文件",
            filetypes=[("支持的格式", "*.zip *.xlsx"), ("ZIP文件", "*.zip"),
                       ("Excel文件", "*.xlsx"), ("所有文件", "*.*")])
        if p:
            self.zip_path.set(p)

    def _select_consult_folder(self):
        p = filedialog.askdirectory(title="选择咨询文件所在文件夹")
        if p:
            self.zip_path.set(p)

    def _select_log_file(self):
        p = filedialog.askopenfilename(title="选择聊天记录文件", filetypes=[("LOG文件", "*.log"), ("所有文件", "*.*")])
        if p:
            self.log_path.set(p)

    def _select_log_folder(self):
        p = filedialog.askdirectory(title="选择聊天记录所在文件夹")
        if p:
            self.log_path.set(p)

    def _iter_supported_files(self, path, exts):
        if os.path.isfile(path):
            if os.path.splitext(path)[1].lower() not in exts:
                raise ValueError(f"不支持的文件格式: {os.path.splitext(path)[1]}，支持: {', '.join(exts)}")
            return [path]
        if os.path.isdir(path):
            files = []
            for name in sorted(os.listdir(path)):
                full = os.path.join(path, name)
                if os.path.isfile(full) and os.path.splitext(name)[1].lower() in exts:
                    files.append(full)
            if not files:
                raise FileNotFoundError(f"文件夹中没有找到支持的文件: {', '.join(exts)}")
            return files
        raise FileNotFoundError("请选择有效的文件或文件夹")

    def _get_sku_set(self):
        """从多行文本框解析SKU集合"""
        raw = self.sku_text.get("1.0", "end").strip()
        if not raw:
            return None
        skus = set()
        for line in raw.splitlines():
            for s in re.split(r"[,，\s]+", line):
                s = s.strip().lstrip("'")
                if s:
                    skus.add(s)
        return skus if skus else None

    def _get_purchase_keywords(self):
        """从成交话术文本框解析关键词列表"""
        raw = self.keyword_text.get("1.0", "end").strip()
        if not raw:
            return [PURCHASE_KEYWORD_DEFAULT]
        keywords = [line.strip() for line in raw.splitlines() if line.strip()]
        return keywords if keywords else [PURCHASE_KEYWORD_DEFAULT]

    # ───────────────── 配置 ─────────────────
    def _load_config(self):
        cfg = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            cfg.read(CONFIG_FILE, encoding="utf-8")
            if "feishu" in cfg:
                self.feishu_app_id.set(cfg["feishu"].get("app_id", ""))
                self.feishu_app_secret.set(cfg["feishu"].get("app_secret", ""))
                self.feishu_link.set(cfg["feishu"].get("link", ""))
                self.feishu_table_name.set(cfg["feishu"].get("table_name", "原声分析"))

    def _load_keywords_config(self):
        """UI构建后加载成交话术到文本框"""
        cfg = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            cfg.read(CONFIG_FILE, encoding="utf-8")
            if "settings" in cfg:
                kw = cfg["settings"].get("purchase_keywords", "")
                if kw:
                    self.keyword_text.delete("1.0", "end")
                    self.keyword_text.insert("1.0", kw.replace("||", "\n"))

    def _on_save_config(self):
        cfg = configparser.ConfigParser()
        cfg["feishu"] = {
            "app_id": self.feishu_app_id.get(),
            "app_secret": self.feishu_app_secret.get(),
            "link": self.feishu_link.get(),
            "table_name": self.feishu_table_name.get().strip() or "原声分析",
        }
        # 成交话术用 || 分隔存储（避免换行符问题）
        kw_raw = self.keyword_text.get("1.0", "end").strip()
        kw_lines = [line.strip() for line in kw_raw.splitlines() if line.strip()]
        cfg["settings"] = {
            "purchase_keywords": "||".join(kw_lines),
        }
        os.makedirs(APP_DATA_DIR, exist_ok=True)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            cfg.write(f)
        self._log("配置已保存")
        messagebox.showinfo("提示", "配置已保存")

    # ───────────────── 飞书 API ─────────────────
    def _get_tenant_token(self):
        url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
        resp = requests.post(url, json={
            "app_id": self.feishu_app_id.get().strip(),
            "app_secret": self.feishu_app_secret.get().strip(),
        }, timeout=15)
        data = resp.json()
        if data.get("code") == 0:
            return data["tenant_access_token"]
        raise Exception(f"获取tenant_access_token失败: {data.get('msg', '未知错误')}")

    def _parse_feishu_link(self):
        link = self.feishu_link.get().strip()
        parsed = urlparse(link)
        path_parts = [p for p in parsed.path.strip("/").split("/") if p]
        app_token = path_parts[-1] if path_parts else ""
        qs = parse_qs(parsed.query)
        table_id = qs.get("table", [""])[0]
        if not app_token:
            raise ValueError("无法解析链接中的app_token")
        return app_token, table_id

    def _on_test_feishu(self):
        try:
            token = self._get_tenant_token()
            app_token, table_id = self._parse_feishu_link()
            self._log(f"Token获取成功，app_token={app_token}, table_id={table_id}")
            if table_id:
                url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields"
                resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=15)
                data = resp.json()
                if data.get("code") == 0:
                    fields = [f["field_name"] for f in data["data"]["items"]]
                    self._log(f"表格字段: {', '.join(fields)}")
                    messagebox.showinfo("成功", f"连接成功!\n表格字段:\n{chr(10).join(fields)}")
                else:
                    self._log(f"读取表格失败: {data.get('msg')}")
                    messagebox.showwarning("警告", f"Token获取成功，但读取表格失败:\n{data.get('msg')}")
            else:
                self._log("链接中无table参数，将在写入时新建表格")
                messagebox.showinfo("成功", "Token获取成功!\n链接中未含table_id，写入时将新建表格")
        except Exception as e:
            self._log(f"测试失败: {e}")
            messagebox.showerror("错误", str(e))

    def _feishu_create_table(self, token, app_token, table_name):
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables"
        body = {
            "table": {
                "name": table_name,
                "default_view_name": "默认视图",
                "fields": [
                    {"field_name": c, "type": 1}
                    for c in OUTPUT_COLUMNS
                ],
            }
        }
        for f in body["table"]["fields"]:
            if f["field_name"] == "进线次数":
                f["type"] = 2
                f["property"] = {"formatter": "0"}
        resp = requests.post(url,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=body, timeout=30)
        data = resp.json()
        if data.get("code") == 0:
            new_table_id = data["data"]["table_id"]
            self._log(f"新建表格成功: {new_table_id}")
            return new_table_id
        raise Exception(f"新建表格失败: {data.get('msg')}")

    def _feishu_ensure_integer_field(self, token, app_token, table_id):
        url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields"
        resp = requests.get(url,
            headers={"Authorization": f"Bearer {token}"},
            timeout=30)
        data = resp.json()
        if data.get("code") != 0:
            self._log(f"读取字段失败，跳过整数格式检查: {data.get('msg')}")
            return {}

        fields = {f.get("field_name"): f for f in data.get("data", {}).get("items", [])}
        line_count_field = fields.get("进线次数")
        if not line_count_field:
            return fields

        if line_count_field.get("type") != 2:
            self._log("进线次数不是数字字段，将按文本写入整数")
            return fields

        prop = line_count_field.get("property") or {}
        if prop.get("formatter") == "0":
            return fields

        field_id = line_count_field.get("field_id")
        if not field_id:
            return fields

        update_url = f"{url}/{field_id}"
        body = {
            "field_name": "进线次数",
            "type": 2,
            "property": {"formatter": "0"},
        }
        resp = requests.put(update_url,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=body, timeout=30)
        data = resp.json()
        if data.get("code") == 0:
            self._log("已将进线次数设置为整数格式")
            line_count_field["property"] = {"formatter": "0"}
        else:
            self._log(f"进线次数整数格式设置失败，继续写入: {data.get('msg')}")
        return fields

    # ───────────────── 数据解析 ─────────────────
    def _parse_consult_input(self):
        zp = self.zip_path.get().strip()
        if not zp or not os.path.exists(zp):
            raise FileNotFoundError("请先选择有效的咨询文件或文件夹")

        files = self._iter_supported_files(zp, CONSULT_EXTS)
        all_rows = []
        for fp in files:
            ext = os.path.splitext(fp)[1].lower()
            self._log(f"读取咨询数据: {os.path.basename(fp)}")
            if ext == ".xlsx":
                rows = self._parse_from_xlsx(fp)
            elif ext == ".zip":
                rows = self._parse_from_zip(fp)
            else:
                continue
            all_rows.extend(rows)

        self._log(f"咨询数据合计: {len(files)} 个文件，{len(all_rows)} 条")
        return all_rows

    def _parse_from_zip(self, zp):
        """从ZIP中提取CSV并解析"""
        with zipfile.ZipFile(zp) as z:
            names = [n for n in z.namelist()
                     if not n.endswith("/") and os.path.splitext(n)[1].lower() == ".csv"]
            if not names:
                raise ValueError("ZIP中没有找到CSV文件")
            csv_name = names[0]
            raw = z.read(csv_name)

        # 尝试多种编码
        text = None
        for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"):
            try:
                text = raw.decode(enc)
                self._log(f"[ZIP→CSV] 编码: {enc}")
                break
            except (UnicodeDecodeError, LookupError):
                continue
        if text is None:
            raise ValueError("无法识别CSV文件编码")
        text = text.lstrip("\ufeff")
        if not text.strip():
            raise ValueError("CSV文件为空")

        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        self._log(f"[ZIP→CSV] 列数: {len(headers)}")

        rows = []
        for src in reader:
            row = {}
            for h in headers:
                val = src.get(h, "")
                row[h] = str(val or "").strip().lstrip("'").strip()
            raw_time = row.get("咨询时间", "")
            m = re.match(r"(\d{4}-\d{2}-\d{2})", raw_time)
            row["咨询日期"] = m.group(1) if m else raw_time
            rows.append(row)

        self._log(f"CSV 解析完成: {len(rows)} 条")
        return rows

    def _parse_from_xlsx(self, fp):
        """直接从Excel文件解析"""
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb.active

        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            wb.close()
            raise ValueError("Excel文件为空")

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        self._log(f"[Excel] 列数: {len(headers)}")

        rows = []
        for xrow in row_iter:
            row = {}
            for j, h in enumerate(headers):
                if not h:
                    continue
                val = xrow[j] if j < len(xrow) else None
                if val is None:
                    val = ""
                elif isinstance(val, float):
                    # 去掉浮点数的 .0 后缀（如SKU编号 10201807261697.0 → 10201807261697）
                    val = str(int(val)) if val == int(val) else str(val)
                else:
                    val = str(val)
                val = val.strip().lstrip("'").strip()
                row[h] = val
            raw_time = row.get("咨询时间", "")
            m = re.match(r"(\d{4}-\d{2}-\d{2})", raw_time)
            row["咨询日期"] = m.group(1) if m else raw_time
            rows.append(row)

        wb.close()
        self._log(f"Excel 解析完成: {len(rows)} 条")
        return rows

    _RE_USER_LINE = re.compile(r"^(.+?)\s+(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s*$")

    def _parse_chat_log(self, purchase_keywords=None):
        if not purchase_keywords:
            purchase_keywords = [PURCHASE_KEYWORD_DEFAULT]
        lp = self.log_path.get().strip()
        if not lp or not os.path.exists(lp):
            raise FileNotFoundError("请先选择有效的LOG文件或文件夹")

        raw_lines = []
        log_files = self._iter_supported_files(lp, LOG_EXTS)
        for fp in log_files:
            self._log(f"读取聊天记录: {os.path.basename(fp)}")
            with open(fp, "r", encoding="utf-8", errors="replace") as f:
                raw_lines.extend(f.readlines())
                raw_lines.append("\n")

        sessions_raw = []
        cur = []
        in_session = False

        for line in raw_lines:
            s = line.rstrip("\n").rstrip("\r")
            if "以下为一通会话" in s:
                in_session = True
                cur = [s]
            elif "会话结束" in s:
                if in_session:
                    cur.append(s)
                    sessions_raw.append(cur)
                cur = []
                in_session = False
            elif in_session:
                cur.append(s)

        self._log(f"聊天记录会话数: {len(sessions_raw)}")

        customer_sessions = defaultdict(list)

        for ses_lines in sessions_raw:
            customer_id = None
            session_dt = None
            cust_msgs = []
            agent_msgs = []
            dialogue_lines = []  # 分行对话格式

            i = 0
            while i < len(ses_lines):
                stripped = ses_lines[i].rstrip("\t").strip()
                m = self._RE_USER_LINE.match(stripped)
                if m:
                    username = m.group(1)
                    timestamp = m.group(2)
                    msg = ""
                    if i + 1 < len(ses_lines):
                        msg = ses_lines[i + 1].rstrip("\t").strip().strip("\ufeff")
                    is_agent = ("图拉斯" in username) or username.startswith("jimi_vender")
                    # 构建对话行
                    if msg:
                        dialogue_lines.append(f"{username} {timestamp}")
                        dialogue_lines.append(msg)
                    if not is_agent:
                        if customer_id is None:
                            customer_id = username
                            session_dt = timestamp
                        if username == customer_id and msg:
                            cust_msgs.append(msg)
                    else:
                        if msg:
                            agent_msgs.append(msg)
                    i += 2
                else:
                    i += 1

            if customer_id:
                full_text = "\n".join(dialogue_lines)
                # 用纯文本做关键词匹配
                flat_text = " ".join(dialogue_lines)
                has_purchase = any(kw in flat_text for kw in purchase_keywords)
                customer_sessions[customer_id].append({
                    "datetime": session_dt,
                    "text": full_text,
                    "customer_messages": cust_msgs,
                    "agent_messages": agent_msgs,
                    "has_purchase": has_purchase,
                })

        self._log(f"解析到 {len(customer_sessions)} 个不同客户")
        return customer_sessions

    def _match_and_build(self, csv_rows, cust_sessions, sku_set=None):
        for cid in cust_sessions:
            cust_sessions[cid].sort(key=lambda s: s["datetime"] or "")

        # 预收集每个客户的所有SKU（按CSV行顺序）
        cust_skus = defaultdict(list)
        for row in csv_rows:
            cid = row.get("顾客", "")
            sku = row.get("商品编号", "")
            if sku and sku not in cust_skus[cid]:
                cust_skus[cid].append(sku)

        results = []
        matched_count = 0
        skipped_sku = 0

        # —— 按客户ID分组CSV行，同一客户合并为一行输出 ——
        cust_csv_group = OrderedDict()  # cid -> list[row]
        for row in csv_rows:
            product_id = row.get("商品编号", "")
            if sku_set and product_id not in sku_set:
                skipped_sku += 1
                continue
            cid = row.get("顾客", "")
            cust_csv_group.setdefault(cid, []).append(row)

        for cid, rows in cust_csv_group.items():
            sessions = cust_sessions.get(cid, [])
            total_sessions = len(sessions)

            # 合并接待客服（去重，保持顺序）
            agents_seen = []
            for row in rows:
                a = row.get("客服", "")
                if a and a not in agents_seen:
                    agents_seen.append(a)
            agent_display = "、".join(agents_seen)

            # 取最早的咨询日期
            dates = [row.get("咨询日期", "") for row in rows]
            dates = [d for d in dates if d]
            consult_date = min(dates) if dates else ""

            # 如果没有匹配的聊天会话，跳过
            if not sessions:
                continue

            has_purchase = any(s["has_purchase"] for s in sessions)

            # 检查是否至少有一个会话有内容
            any_text = any(s["text"].strip() for s in sessions)
            if not any_text:
                continue

            matched_count += 1
            separator = "\n————————————————\n"

            if len(sessions) > 1:
                # 多次进线：所有进线都用分割线分隔
                chat_parts = []
                cust_parts = []
                agent_parts = []
                for idx, s in enumerate(sessions):
                    tag = f"【第{idx + 1}次进线】"
                    chat_parts.append(f"{tag}\n{s['text']}")
                    c_msgs = [m for m in s["customer_messages"]
                              if not _RE_URL.search(m)]
                    if c_msgs:
                        c_text = "\n".join(f"{i + 1}、{m}" for i, m in enumerate(c_msgs))
                        cust_parts.append(f"{tag}\n{c_text}")
                    a_msgs = [m for m in s["agent_messages"]
                              if not _RE_URL.search(m)]
                    if a_msgs:
                        a_text = "\n".join(f"{i + 1}、{m}" for i, m in enumerate(a_msgs))
                        agent_parts.append(f"{tag}\n{a_text}")

                chat_text_all = separator.join(chat_parts)
                cust_info = separator.join(cust_parts)
                agent_reply = separator.join(agent_parts)
                all_skus = cust_skus.get(cid, [])
                if len(all_skus) > 1:
                    sku_display = separator.join(f"【SKU {i + 1}】{s}" for i, s in enumerate(all_skus))
                else:
                    sku_display = all_skus[0] if all_skus else rows[0].get("商品编号", "")
            else:
                # 单次进线
                s = sessions[0]
                chat_text_all = s["text"]
                c_msgs = [m for m in s["customer_messages"]
                          if not _RE_URL.search(m)]
                cust_info = "\n".join(f"{i + 1}、{m}" for i, m in enumerate(c_msgs))
                a_msgs = [m for m in s["agent_messages"]
                          if not _RE_URL.search(m)]
                agent_reply = "\n".join(f"{i + 1}、{m}" for i, m in enumerate(a_msgs))
                all_skus = cust_skus.get(cid, [])
                sku_display = all_skus[0] if all_skus else rows[0].get("商品编号", "")

            results.append({
                "咨询时间": consult_date,
                "客户ID": cid,
                "接待客服": agent_display,
                "商品编号": sku_display,
                "进线次数": total_sessions,
                "是否购买": "是" if has_purchase else "否",
                "聊天记录": chat_text_all,
                "客户问题原话": cust_info,
                "客服回复": agent_reply,
            })

        if sku_set:
            self._log(f"SKU过滤: 跳过 {skipped_sku} 条，保留 {len(results)} 条")
        self._log(f"生成 {len(results)} 条结果，匹配聊天记录 {matched_count} 条")
        return results

    # ───────────────── 按钮回调 ─────────────────
    def _on_parse(self):
        def task():
            try:
                self._log("--- 开始解析 ---")
                self.progress["value"] = 0
                self._update_stats()

                csv_rows = self._parse_consult_input()
                self.csv_data = csv_rows
                self.progress["value"] = 30

                purchase_keywords = self._get_purchase_keywords()
                self._log(f"成交话术关键词: {', '.join(purchase_keywords)}")
                cust_sessions = self._parse_chat_log(purchase_keywords)
                self.chat_sessions = cust_sessions
                self.progress["value"] = 70

                sku_set = self._get_sku_set()
                if sku_set:
                    self._log(f"SKU筛选: {', '.join(sku_set)}")

                self.result_data = self._match_and_build(csv_rows, cust_sessions, sku_set)
                self.progress["value"] = 100
                self._update_stats(len(csv_rows), sum(len(v) for v in cust_sessions.values()), len(self.result_data))
                self._log(f"--- 解析完成！共 {len(self.result_data)} 条 ---")
                messagebox.showinfo("完成",
                    f"解析完成!\n"
                    f"CSV记录: {len(csv_rows)}\n"
                    f"聊天会话: {sum(len(v) for v in cust_sessions.values())}\n"
                    f"输出行数: {len(self.result_data)}")
            except Exception as e:
                self._log(f"解析错误: {e}")
                messagebox.showerror("错误", str(e))
            finally:
                self._finish_task()

        self._start_task("正在解析数据...", task)

    def _on_write_feishu(self):
        if not self.result_data:
            messagebox.showwarning("警告", "请先点击「解析」")
            return

        def task():
            try:
                token = self._get_tenant_token()
                app_token, table_id = self._parse_feishu_link()

                if self.feishu_create_new.get():
                    table_name = self.feishu_table_name.get().strip() or "原声分析"
                    self._log(f"新建表格: {table_name}")
                    table_id = self._feishu_create_table(token, app_token, table_name)
                else:
                    if not table_id:
                        raise ValueError("链接中没有table参数，请勾选「新建表格」或在链接中包含table参数")
                    self._log(f"写入现有表格: {table_id}")

                data = self.result_data
                self._log(f"开始写入飞书 (共 {len(data)} 条)...")
                self.progress["value"] = 0
                feishu_fields = self._feishu_ensure_integer_field(token, app_token, table_id)
                line_count_is_number = (feishu_fields.get("进线次数") or {}).get("type") == 2

                batch_size = 100
                total = len(data)

                for start in range(0, total, batch_size):
                    batch = data[start:start + batch_size]
                    records = []
                    for r in batch:
                        fields = {}
                        for col in OUTPUT_COLUMNS:
                            val = r[col]
                            if col == "进线次数":
                                val = int(float(val or 0))
                                if not line_count_is_number:
                                    val = str(val)
                            elif col in ("聊天记录", "客户问题原话", "客服回复", "商品编号"):
                                val = str(val)[:50000]
                            fields[col] = val
                        records.append({"fields": fields})

                    url = (f"https://open.feishu.cn/open-apis/bitable/v1/apps/"
                           f"{app_token}/tables/{table_id}/records/batch_create")
                    resp = requests.post(url,
                        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                        json={"records": records},
                        timeout=60)
                    rdata = resp.json()
                    if rdata.get("code") != 0:
                        raise Exception(f"写入失败(batch {start}): {rdata.get('msg')}")

                    done = min(start + batch_size, total)
                    self.progress["value"] = int(done / total * 100)
                    self._log(f"已写入 {done}/{total}")

                self.progress["value"] = 100
                self._log("--- 写入飞书完成 ---")
                messagebox.showinfo("完成", f"成功写入 {total} 条记录到飞书")
            except Exception as e:
                self._log(f"飞书写入错误: {e}")
                messagebox.showerror("错误", str(e))
            finally:
                self._finish_task()

        self._start_task("正在写入飞书...", task)

    def _on_export_excel(self):
        if not self.result_data:
            messagebox.showwarning("警告", "请先点击「解析」")
            return

        save_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            initialfile="客服分析结果.xlsx",
            filetypes=[("Excel文件", "*.xlsx")])
        if not save_path:
            return

        def task():
            try:
                self._log("生成Excel中...")
                self.progress["value"] = 0

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "客服分析"

                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", name="Microsoft YaHei UI", size=10)
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

                for col_idx, h in enumerate(OUTPUT_COLUMNS, 1):
                    c = ws.cell(row=1, column=col_idx, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = thin_border

                data = self.result_data
                total = len(data)
                for i, r in enumerate(data, 2):
                    longest_lines = 1
                    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
                        val = r[col_name]
                        if col_name in ("聊天记录", "客户问题原话", "客服回复", "商品编号"):
                            val = str(val)
                            if len(val) > 32700:
                                val = val[:32700] + "...(已截断)"
                            longest_lines = max(longest_lines, min(val.count("\n") + 1, 12))
                        cell = ws.cell(row=i, column=col_idx, value=val)
                        cell.border = thin_border
                        if col_name in ("聊天记录", "客户问题原话", "客服回复", "商品编号"):
                            cell.alignment = Alignment(wrap_text=True, vertical="top")

                    ws.row_dimensions[i].height = max(18, min(120, longest_lines * 15))

                    if (i - 1) % 500 == 0:
                        self.progress["value"] = int((i - 1) / total * 90)

                # A~I 列宽
                widths = {"A": 13, "B": 22, "C": 16, "D": 18, "E": 10, "F": 10, "G": 60, "H": 45, "I": 45}
                for col_letter, w in widths.items():
                    ws.column_dimensions[col_letter].width = w

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                wb.save(save_path)
                self.last_output_dir = os.path.dirname(os.path.abspath(save_path))
                self.progress["value"] = 100
                self._log(f"Excel已保存: {save_path}")
                messagebox.showinfo("完成", f"Excel已保存到:\n{save_path}")
            except Exception as e:
                self._log(f"Excel错误: {e}")
                messagebox.showerror("错误", str(e))
            finally:
                self._finish_task()

        self._start_task("正在生成Excel...", task)


def main():
    root = tk.Tk()
    ChatAnalyzerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
